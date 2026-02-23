from datetime import datetime, timedelta
from pathlib import Path
from threading import Thread
from uuid import uuid4

from fastapi import BackgroundTasks, Depends, FastAPI, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .config import settings
from .crud import (
    bulk_insert_users,
    create_upload_job,
    get_upload_jobs_by_status,
    get_jobs_for_cleanup,
    get_upload_job,
    mark_upload_file_deleted,
    set_upload_job_completed,
    set_upload_job_failed,
    set_upload_job_running,
)
from .database import Base, SessionLocal, engine
from .models import User

Base.metadata.create_all(bind=engine)

app = FastAPI()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _save_upload_file(file: UploadFile) -> Path:
    upload_dir = Path(settings.UPLOAD_DIR)
    upload_dir.mkdir(parents=True, exist_ok=True)

    extension = Path(file.filename or "upload.xlsx").suffix or ".xlsx"
    unique_name = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex}{extension}"
    file_path = upload_dir / unique_name

    with file_path.open("wb") as out_file:
        while True:
            chunk = file.file.read(1024 * 1024)
            if not chunk:
                break
            out_file.write(chunk)

    return file_path


def _normalize_row(raw_row: tuple, column_map: dict[str, int]) -> dict:
    def get_value(column_name: str):
        return raw_row[column_map[column_name]]

    name = str(get_value("name") or "").strip()
    email = str(get_value("email") or "").strip()

    age_value = get_value("age")
    try:
        age = int(float(age_value)) if age_value not in (None, "") else 0
    except (TypeError, ValueError):
        age = 0

    return {"name": name, "email": email, "age": age}


def _serialize_job(job) -> dict:
    return {
        "job_id": job.job_id,
        "status": job.status,
        "filename": job.filename,
        "saved_path": job.saved_path,
        "chunk_size": job.chunk_size,
        "inserted_rows": job.inserted_rows,
        "error": job.error,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "started_at": job.started_at.isoformat() if job.started_at else None,
        "completed_at": job.completed_at.isoformat() if job.completed_at else None,
        "file_deleted_at": job.file_deleted_at.isoformat() if job.file_deleted_at else None,
    }


def cleanup_expired_uploads(db: Session) -> int:
    if not settings.CLEANUP_EXPIRED_UPLOADS:
        return 0

    cutoff = datetime.utcnow() - timedelta(hours=settings.UPLOAD_FILE_RETENTION_HOURS)
    cleanup_jobs = get_jobs_for_cleanup(db, cutoff)

    deleted_files = 0
    for job in cleanup_jobs:
        try:
            file_path = Path(job.saved_path)
            if file_path.exists() and file_path.is_file():
                file_path.unlink()
                deleted_files += 1
            mark_upload_file_deleted(db, job.job_id)
        except OSError:
            continue

    return deleted_files


def process_excel_job(job_id: str, file_path: Path) -> None:
    db = SessionLocal()
    workbook = None

    try:
        set_upload_job_running(db, job_id)

        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        worksheet = workbook.active

        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            raise ValueError("Excel file is empty")

        header_map = {
            str(value).strip().lower(): index
            for index, value in enumerate(headers)
            if value is not None and str(value).strip()
        }

        missing_columns = [
            column for column in settings.REQUIRED_COLUMNS if column not in header_map
        ]
        if missing_columns:
            raise ValueError(
                f"Excel must contain columns: {settings.REQUIRED_COLUMNS}. "
                f"Missing: {missing_columns}"
            )

        column_map = {
            "name": header_map["name"],
            "email": header_map["email"],
            "age": header_map["age"],
        }

        batch: list[dict] = []
        inserted_rows = 0

        for row in rows:
            if row is None:
                continue

            user = _normalize_row(row, column_map)

            if not user["name"] and not user["email"] and user["age"] == 0:
                continue

            batch.append(user)

            if len(batch) >= settings.EXCEL_CHUNK_SIZE:
                bulk_insert_users(db, batch)
                inserted_rows += len(batch)
                batch.clear()

        if batch:
            bulk_insert_users(db, batch)
            inserted_rows += len(batch)

        set_upload_job_completed(db, job_id, inserted_rows)

    except Exception as exc:
        set_upload_job_failed(db, job_id, str(exc))
    finally:
        if workbook is not None:
            workbook.close()
        db.close()


def _recover_jobs_after_restart() -> None:
    db = SessionLocal()
    try:
        interrupted_jobs = get_upload_jobs_by_status(db, ["running"])
        for job in interrupted_jobs:
            set_upload_job_failed(
                db,
                job.job_id,
                "Job interrupted because server restarted",
            )

        queued_jobs = get_upload_jobs_by_status(db, ["queued"])
    finally:
        db.close()

    for job in queued_jobs:
        file_path = Path(job.saved_path)
        if not file_path.exists():
            db_missing = SessionLocal()
            try:
                set_upload_job_failed(
                    db_missing,
                    job.job_id,
                    "Queued job file not found after server restart",
                )
            finally:
                db_missing.close()
            continue

        Thread(
            target=process_excel_job,
            args=(job.job_id, file_path),
            daemon=True,
        ).start()


@app.on_event("startup")
def startup_recovery() -> None:
    _recover_jobs_after_restart()


@app.get("/")
def read_root():
    return {"message": "API is running"}


@app.get("/count")
def count_users(db: Session = Depends(get_db)):
    return {"count": db.query(User).count()}


@app.post("/upload-excel/")
async def upload_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Uploaded file name is missing")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    cleanup_expired_uploads(db)

    file_path = _save_upload_file(file)
    job_id = uuid4().hex

    create_upload_job(
        db=db,
        job_id=job_id,
        filename=file.filename,
        saved_path=str(file_path),
        chunk_size=settings.EXCEL_CHUNK_SIZE,
    )

    background_tasks.add_task(process_excel_job, job_id, file_path)

    return {
        "message": "File saved and processing job created",
        "job_id": job_id,
        "status": "queued",
        "saved_path": str(file_path),
        "chunk_size": settings.EXCEL_CHUNK_SIZE,
    }


@app.get("/jobs/{job_id}")
def get_job_status(job_id: str, db: Session = Depends(get_db)):
    cleanup_expired_uploads(db)

    job = get_upload_job(db, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    return _serialize_job(job)


@app.post("/jobs/cleanup")
def cleanup_jobs(db: Session = Depends(get_db)):
    deleted_files = cleanup_expired_uploads(db)
    return {
        "message": "Cleanup completed",
        "deleted_files": deleted_files,
        "retention_hours": settings.UPLOAD_FILE_RETENTION_HOURS,
    }